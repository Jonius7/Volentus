import json
from openpyxl import load_workbook

# Columns to ignore completely
ignore_columns = {"License", "Filename", "Notes", "Repo", "sourceFileName", "followsIL", "CF URL"}

# Explicit mapping of nested columns
nested_columns = {
    "installationPolicy": ["continueOnFailedDownload", "optionalKey", "selectedByDefault", "description", "name"],
    "metadata": ["side"]
}
list_columns = {"follows"}

# Sheet name → top-level key in JSON
sheet_mapping = {
    "CurseForge": "curse",
    "URL": "url",
    "Modrinth": "modrinth",
    "Core(CF)": "curse",
    "Core(URL)": "url",
    "Core(MR)": "modrinth",
    "MediaFire": "url"
}

# Load workbook
print("Loading workbook...")
wb = load_workbook("bundle.json v5.xlsx")

mods_bundle = {}
coremods_bundle = {}
optional_bundle = {}

def process_sheet(sheet, nested_columns, ignore_columns, list_columns=None):
    if list_columns is None:
        list_columns = set()

    headers = [cell.value for cell in sheet[1]]
    items = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        if not any(row_dict.values()):
            continue

        obj = {}

        # Build nested objects
        for parent, children in nested_columns.items():
            nested_obj = {}
            for child in children:
                value = row_dict.get(child)
                if value not in (None, ""):
                    if child in list_columns:
                        nested_obj[child] = [value]  # wrap as list
                    else:
                        nested_obj[child] = value
            if nested_obj:
                obj[parent] = nested_obj

        # Top-level fields (not nested, not ignored)
        for key, value in row_dict.items():
            if (
                key not in ignore_columns
                and key not in sum(nested_columns.values(), [])
                and value not in (None, "")
            ):
                if key in list_columns:
                    obj[key] = [value]  # wrap as list
                else:
                    obj[key] = value

        if obj:
            items.append(obj)

    return items


# Only process mapped sheets
mapped_sheets = [s for s in wb.sheetnames if s in sheet_mapping]
print("Mapped sheets found:", mapped_sheets)

for idx, sheet_name in enumerate(mapped_sheets):
    print(f"Processing sheet: {sheet_name}")
    sheet = wb[sheet_name]
    processed_items = process_sheet(sheet, nested_columns, ignore_columns, list_columns)
    top_key = sheet_mapping[sheet_name]

    if idx < 3:
        mods_bundle[top_key] = processed_items
    elif idx >= 3 and idx < 6:
        coremods_bundle[top_key] = processed_items
    else:
        optional_bundle[top_key] = processed_items
        

# Save JSON bundles
if mods_bundle:
    with open("mods.bundle.json", "w") as f:
        json.dump(mods_bundle, f, indent=4)
    print("Saved mods.bundle.json")

if coremods_bundle:
    with open("coremods.bundle.json", "w") as f:
        json.dump(coremods_bundle, f, indent=4)
    print("Saved coremods.bundle.json")

if optional_bundle:
    with open("optional.bundle.json", "w") as f:
        json.dump(optional_bundle, f, indent=4)
    print("Saved optional.bundle.json")